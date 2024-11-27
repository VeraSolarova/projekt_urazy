from openpyxl import Workbook, load_workbook
wb = load_workbook ('scripts/vek.xlsx')

print(wb.sheetnames)
